from pypinyin import pinyin, Style
import os
import string
import random
import pygame
from app.config.app_config import globalAppSettings, Queue_lock
from app.utils.asr.rtasr_python3_demo import XF_Client
from app.utils.sound.recorder import Recorder
from mutagen import File
from mutagen.mp3 import MP3
from mutagen.mp4 import MP4
from mutagen.wave import WAVE
from datetime import datetime, timedelta
from pydub import AudioSegment
import wave
import pyaudio
import time
import shutil
import pandas as pd
import openpyxl

def chinese_to_pinyin(chinese_str):  
    """  
    将中文字符串转换为拼音。  
    """  
    # Style.NORMAL表示返回普通拼音，不带声调  
    pinyin_list = pinyin(chinese_str, style=Style.NORMAL)  
      
    # 将拼音列表转换为字符串，拼音之间用空格分隔
    pinyin_str = ''.join([''.join(item) for item in pinyin_list])  
      
    return pinyin_str  

def get_file_dir():
    if not os.path.exists(globalAppSettings.audio_path):
        os.makedirs(globalAppSettings.audio_path)
    return globalAppSettings.audio_path

def get_mic_dir():
    if not os.path.exists(globalAppSettings.mic_path):  
        os.makedirs(globalAppSettings.mic_path)
    return globalAppSettings.mic_path

def get_pic_dir():
    if not os.path.exists(globalAppSettings.photo_dir):  
        os.makedirs(globalAppSettings.photo_dir)
    return globalAppSettings.photo_dir

def get_excel_dir():
    if not os.path.exists(globalAppSettings.excel_dir):  
        os.makedirs(globalAppSettings.excel_dir)
    return globalAppSettings.excel_dir

def get_audio_duration(file_path):
    """
    获取音频文件的时长（秒），并处理可能的错误。
    """
    if not os.path.exists(file_path):
        return {"status": "error", "data": f"Error: {file_path} not found."}
    
    file_extension = os.path.splitext(file_path)[1].lower()
    time.sleep(1)
    # 尝试获取音频时长，最多重新尝试一次
    for attempt in range(2):
        try:
            if file_extension == '.mp3':
                audio = MP3(file_path)
            elif file_extension == '.mp4':
                audio = MP4(file_path)
            elif file_extension == '.wav':
                audio = WAVE(file_path)
            else:
                return {"status": "error", "data": f"Unsupported audio format: {file_extension}"}
            
            # 获取时长（秒），并精确到小数点后1位
            duration = audio.info.length
            rounded_duration = round(duration, 1)
            return {"status": "ok", "data": str(rounded_duration)}
        
        except Exception as e:
            if attempt == 0:
                print('fail1', e)
                # 第一次尝试失败，等待1秒后重新尝试
                time.sleep(1)
            else:
                # 第二次尝试失败，返回错误信息
                print('fail2', e)
                return {"status": "error", "data": f"An unexpected error occurred: {e}"}

def get_audio_duration_ms(file_path):
    """  
    获取音频文件的时长（ms），并处理可能的错误。  
    """
    if not os.path.exists(file_path):
        return "error"
    file_extension = os.path.splitext(file_path)[1].lower()
    try:  
        if file_extension == '.mp3':
            audio = MP3(file_path)
        elif file_extension == '.mp4':
            audio = MP4(file_path)
        elif file_extension == '.wav':
            audio = WAVE(file_path)
        else:
            return "error"
        # 获取时长（毫秒），然后转换为秒  
        duration = audio.info.length  # 获取音频时长（秒）
        duration_ms = round(duration * 1000)
        return duration_ms
    except Exception as e:  
        # 捕获并处理文件损坏或解析错误的情况
        return "error"

def generate_random_string(length=6):
    # 字母和数字的组合
    characters = string.ascii_letters + string.digits
    base = len(characters)
    
    # 获取当前时间戳
    timestamp = int(time.time() * 1000)  # 毫秒级时间戳
    
    # 将时间戳转换为指定字符集合的编码字符串
    encoded_string = []
    while timestamp > 0:
        timestamp, remainder = divmod(timestamp, base)
        encoded_string.append(characters[remainder])
    
    # 将编码结果反转并补全到所需长度
    encoded_string = ''.join(reversed(encoded_string))
    if len(encoded_string) < length:
        encoded_string = characters[-1] * (length - len(encoded_string)) + encoded_string

    # 反转字符顺序，使其字典顺序减小
    reversed_characters = characters[::-1]
    descending_string = ''.join(reversed_characters[characters.index(c)] for c in encoded_string)
    
    return descending_string[-length:]  # 确保长度固定

def play_audio_with_pygame(file_path):
    # ile_path = get_file_dir() + '/' + file_name
    pygame.mixer.init()  
    try:  
        pygame.mixer.music.load(file_path)  
        pygame.mixer.music.play()  
        while pygame.mixer.music.get_busy():  
            pygame.time.Clock().tick(10)  
        # print("音频播放完毕。")  
    except Exception as e:  
        print(f"播放音频时出错: {e}")  
    finally:  
        pygame.quit()

def get_beijing_time():
    return datetime.utcnow() + timedelta(hours=8)

def create_audio_client():
    audio = pyaudio.PyAudio()
    return audio

def close_audio_client(audio):
    audio.terminate()

def record_audio(audio, mic_file_name):
    FORMAT = pyaudio.paInt16  # 16位音频格式
    CHANNELS = 1  # 立体声
    RATE = 16000  # 采样率
    CHUNK = 1024  # 每个块的帧数
    RECORD_SECONDS = 5  # 录音时间

    stream = audio.open(format=FORMAT, channels=CHANNELS,
                        rate=RATE, input=True,
                        frames_per_buffer=CHUNK)

    frames = []
    for i in range(0, int(RATE / CHUNK * RECORD_SECONDS)):
        data = stream.read(CHUNK)
        frames.append(data)

    stream.stop_stream()
    stream.close()
    # time.sleep(1)
    storage_path = get_mic_dir()
    filepath = os.path.join(storage_path, mic_file_name)
    with open(filepath, 'wb') as f:
        f.write(b''.join(frames))
    # print(f"录音已保存为 {filepath}")
    return filepath

def record_audio_to_wav(audio, wav_name):
    FORMAT = pyaudio.paInt16  # 16位音频格式
    CHANNELS = 1  # 立体声
    RATE = 44100  # 采样率
    CHUNK = 1024  # 每个块的帧数
    RECORD_SECONDS = 5  # 录音时间

    stream = audio.open(format=FORMAT, channels=CHANNELS,
                        rate=RATE, input=True,
                        frames_per_buffer=CHUNK)

    frames = []
    for i in range(0, int(RATE / CHUNK * RECORD_SECONDS)):
        data = stream.read(CHUNK)
        frames.append(data)

    stream.stop_stream()
    stream.close()
    storage_path = get_mic_dir()
    filepath = os.path.join(storage_path, wav_name)
    with wave.open(filepath,"wb") as f:
        f.setnchannels(CHANNELS)
        f.setsampwidth(audio.get_sample_size(FORMAT))
        f.setframerate(RATE)
        f.writeframes(b''.join(frames))
    return filepath

def play_audio_with_pygame_record(q, file_path, audio, filepath):
    # ile_path = get_file_dir() + '/' + file_name
    pygame.mixer.init()
    try:  
        pygame.mixer.music.load(file_path)  
        pygame.mixer.music.play()  
        while pygame.mixer.music.get_busy():  
            pygame.time.Clock().tick(10)

        # micaudio_path = record_audio(audio, mic_file_name)
        recorder = Recorder(output_file=filepath)
        voice_timestamps_list = recorder.run() # 录音
        q.put(voice_timestamps_list)

    except Exception as e:  
        print(f"播放音频时出错: {e}")  
    finally:  
        pygame.quit()

def play_record(q, filepath, audio_time, wait_time, record_time):
    recorder = Recorder(output_file=filepath, first_audio_time=audio_time, chunk_split_silence_duration=wait_time, max_record_time=record_time)
    delay_time = recorder.run() # 录音
    Queue_lock.acquire()
    q.put(delay_time)
    Queue_lock.release()

def audio_to_text_xf(filepath):
# 讯飞语音转文字函数
    client = XF_Client()
    client.send(filepath)
    time.sleep(1)
    res = client.str_output
    client.close()
    return res

def list_all_files(folder_path, file_name=""):
    file_paths = []  # 用于保存文件路径的列表
    for root, _, files in os.walk(folder_path):  # 遍历文件夹
        for file in files:
            if file_name in file:  # 检查文件名是否包含指定字段
                file_path = os.path.join(root, file)  # 构造文件的完整路径
                file_paths.append(file_path)          # 将路径添加到列表中
    return file_paths

def move_file_to_folder(file_path, folder_str):  
    # 获取文件的目录  构造目标文件夹路径
    directory = os.path.dirname(file_path)  
    target_folder = os.path.join(directory, folder_str)  
      
    # 如果目标文件夹不存在，则创建它  
    if not os.path.exists(target_folder):  
        os.makedirs(target_folder)  

    file_name = os.path.basename(file_path)  # 获取文件名  
    target_file_path = os.path.join(target_folder, file_name)  # 构造目标文件路径  
    shutil.move(file_path, target_file_path)# 移动文件  

    return target_file_path

def copy_file_to_modifyfolder(file_path, folder_str):
    # 获取文件的目录  构造目标文件夹路径
    directory = get_file_dir()
    target_folder = os.path.join(directory, folder_str)
      
    # 如果目标文件夹不存在，则创建它  
    if not os.path.exists(target_folder):  
        os.makedirs(target_folder)  

    file_name = os.path.basename(file_path)  # 获取文件名  
    target_file_path = os.path.join(target_folder, file_name)  # 构造目标文件路径  
    shutil.copy(file_path, target_file_path)# 移动文件  

    return target_file_path

def pcm_to_wav(file, rate=16000):
    """
    输入为文件路径和采样频率，默认16k，本系统录音就是用的16k
    输出为这个路径下同名的文件，比如输入路径为
    输入：/Users/dlh/Desktop/zhaoshang/CVAtest/backend/mic_test.pcm
    输出：/Users/dlh/Desktop/zhaoshang/CVAtest/backend/mic_test.wav
    """
    if not os.path.exists(file):
        return ""
    out_name = file.replace('pcm', 'wav')
    with open(file, 'rb') as pcm:
        data = pcm.read()
    with wave.open(out_name, 'wb') as wavfile:
        wavfile.setparams((2, 2, int(rate / 2), 0, 'NONE', 'NONE'))
        wavfile.writeframes(data)
    return out_name

def save_data_to_excel(file_path, file_name, turn_id, configtype, data_list):
    # 尝试生成文件名并检查是否存在
    # file_index = turn_id
    # while True:
    #     full_file_path = os.path.join(file_path, f"{file_name}_{file_index}.xlsx")
    #     if not os.path.exists(full_file_path):
    #         break
    #     file_index += 1
    subname = ""
    if configtype == "rouse":
        subname = "唤醒"
    elif configtype == "false-rouse":
        subname = "误唤醒"
    elif configtype == "interaction":
        subname = "单次对话"
    elif configtype == "interaction-multi":
        subname = "连续对话"
    full_file_path = os.path.join(file_path, f"{file_name}_第{turn_id}轮_{subname}.xlsx")
    if os.path.exists(full_file_path):
        os.remove(full_file_path)

    # 将数据转换为DataFrame格式
    df = pd.DataFrame(data_list)

    # 写入Excel文件
    with pd.ExcelWriter(full_file_path, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name="Sheet1")

    print(f"数据已成功保存到 {full_file_path}")

def judge_voice_language(input1, input2):
    b_value = input1
    voice = -1
    if b_value == "男声":
        voice = 1
    elif b_value == "女声":
        voice = 2

    c_value = input2
    language = -1
    if c_value == "男声1" or c_value == "女声1":
        language = 1
    elif c_value == "男声2" or c_value == "女声2":
        language = 2
    elif c_value == "男声3" or c_value == "女声3":
        language = 3
    elif c_value == "男声4" or c_value == "女声4":
        language = 4
    elif c_value == "男声5" or c_value == "女声5":
        language = 5
    elif c_value == "男声6" or c_value == "女声6":
        language = 6
    elif c_value == "女声7":
        language = 7
    elif c_value == "童声":
        if voice == 1:
            language = 7
        else:
            language = 8
    elif c_value == "东北话":
        if voice == 1:
            language = 8
        else:
            language = 11
    elif c_value == "天津话" or c_value == "四川话":
        language = 9
    elif c_value == "粤语":
        language = 10
    return voice, language

def analysis_excel(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Store the data
    num = 1
    res_list = []
    for row in range(2, sheet.max_row + 1):
        a_value = sheet[f"A{row}"].value

        b_value = sheet[f"B{row}"].value
        c_value = sheet[f"C{row}"].value
        voice, language = judge_voice_language(b_value, c_value)

        d_value = sheet[f"D{row}"].value

        c_type = -1
        if d_value == "测试语料":
            c_type = 1
        elif d_value == "唤醒语料":
            c_type = 2
        elif d_value == "干扰语料":
            c_type = 3

        e_value = sheet[f"E{row}"].value
        f_value = sheet[f"F{row}"].value

        g_value = sheet[f"G{row}"].value
        is_tone = False
        if g_value == "否":
            is_tone = False
        elif g_value == "是":
            is_tone = True

        temp_data = {
            "A": a_value,
            "B": voice,
            "C": language,
            "D": c_type,
            "E": e_value,
            "F": f_value,
            "G": is_tone
        }

        res_list.append(temp_data)
    return res_list

def rouse_excel(sheet):
    # Store the data
    res_list = []
    for row in range(2, sheet.max_row + 1):
        a_value = sheet[f"A{row}"].value
        b_value = sheet[f"B{row}"].value
        test_sc = ""
        if b_value == "唤醒":
            test_sc = "wake-up"
        elif b_value == "声源定位":
            test_sc = "sound-source-localization"
        elif b_value == "误唤醒":
            test_sc = "false-wake-up"
        elif b_value == "车外唤醒":
            test_sc = "external-wake-up"

        c_value = sheet[f"C{row}"].value
        d_value = sheet[f"D{row}"].value

        speaker_tmp, voice = judge_voice_language(c_value, d_value)

        speaker = ""
        if speaker_tmp == 2:
            speaker = "female"
        elif speaker_tmp == 1:
            speaker = "male"
        else:
            speaker = ""

        e_value = sheet[f"E{row}"].value
        f_value = sheet[f"F{row}"].value

        temp_data = {
            "A": a_value,
            "B": test_sc,
            "C": speaker,
            "D": voice,
            "E": e_value,
            "F": f_value
        }

        res_list.append(temp_data)
    return res_list

def test_excel(sheet):
    # Store the data
    res_list = []
    for row in range(2, sheet.max_row + 1):
        a_value = sheet[f"A{row}"].value
        b_value = sheet[f"B{row}"].value
        test_type = ""
        if b_value == "语音识别":
            test_type = "speech-recognition-interaction"
        elif b_value == "智能交互":
            test_type = "intelligent-interaction"
        elif b_value == "免唤醒":
            test_type = "wake-up-free"

        c_value = sheet[f"C{row}"].value
        test_sc = ""
        if c_value == "单指令交互":
            test_sc = "single-command-interaction"
        elif c_value == "连续对话交互":
            test_sc = "continuous-dialogue-interaction"
        elif c_value == "多指令交互":
            test_sc = "multi-command-interaction"
        elif c_value == "模糊指令交互":
            test_sc = "fuzzy-command-interaction"
        elif c_value == "多话题交叉执行":
            test_sc = "multi-topic-cross-execution"

        d_value = sheet[f"D{row}"].value
        e_value = sheet[f"E{row}"].value

        speaker_tmp, voice = judge_voice_language(d_value, e_value)

        speaker = ""
        if speaker_tmp == 2:
            speaker = "female"
        elif speaker_tmp == 1:
            speaker = "male"
        else:
            speaker = ""

        f_value = sheet[f"F{row}"].value
        fun = ""
        if f_value == "音视频":
            fun = "audio-video"
        elif f_value == "导航和出行":
            fun = "navigation-travel"
        elif f_value == "通讯":
            fun = "communication"
        elif f_value == "车辆设置与信息查询":
            fun = "vehicle-settings-info-query"
        elif f_value == "车辆指控指令":
            fun = "vehicle-control-command"
        elif f_value == "AI助手":
            fun = "ai-assistant"
        elif f_value == "安全与隐私":
            fun = "security-privacy"

        g_value = sheet[f"G{row}"].value
        h_value = sheet[f"H{row}"].value
        i_value = sheet[f"I{row}"].value

        temp_data = {
            "A": a_value,
            "B": test_type,
            "C": test_sc,
            "D": speaker,
            "E": voice,
            "F": fun,
            "G": g_value,
            "H": h_value,
            "I": i_value,
        }

        res_list.append(temp_data)
    return res_list

def corpus_excel(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Store the data
    res_list = []
    excel_type = ""
    for row in range(1, 2):
        val = sheet[f"A{row}"].value
        if val == "唤醒语料文本":
            excel_type = "rouse"
            res_list = rouse_excel(sheet)
        else:
            excel_type = "test"
            res_list = test_excel(sheet)
        break
    return excel_type, res_list